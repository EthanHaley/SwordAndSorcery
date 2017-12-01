#Creates a basic queries.sql files
##TO DO: check queries, check formatting of output, add violation handling,
#  create decent party names, produce data from PHB
#  IDEA: move table creation to another file. It will be easier to edit.
#  IDEA #2: change excel files to add consistency to race attributes; e.g. Orc = Size Medium
#  IMPORTANT: This will overwrite any file named 'queries.sql'!

import openpyxl #!!required: pip install openyxl
from random import randint
import random

f = open("queries.sql","w")

#Excel I/O
wb_armor = openpyxl.load_workbook('armor.xlsx')
armor_excel = wb_armor.active
wb_items = openpyxl.load_workbook('items.xlsx')
items_excel = wb_items.active
wb_weapons = openpyxl.load_workbook('weapons.xlsx')
weapons_excel = wb_weapons.active
wb_names_email = openpyxl.load_workbook('first_last_email.xlsx')
names_email_excel = wb_names_email.active
wb_misc = openpyxl.load_workbook('misc.xlsx')
misc_excel = wb_misc.active
wb_spells = openpyxl.load_workbook('spells.xlsx')
spells_excel = wb_spells.active
wb_monsters = openpyxl.load_workbook('monsters.xlsx')
monsters_excel = wb_monsters.active

payment_rates = [0, 5, 10, 15] #Tweak: Change to design. Do we have a free plan?
hit_die = ["1d4", "1d6", "1d8", "1d10", "1d12", "1d20"] #Tweak: possibly unused
races = ["Elf","Dwarf","Human","Halfling","Orc", "Dragonborn", "Tiefling", "Half-Elf", "Half-Orc"] # from PHB --Tweak
classes = ["Paladin","Cleric","Wizard", "Rogue","Ranger", "Monk", "Sorcerer", "Druid", "Barbarian", "Bard"] # from PHB --Tweak
sizes = ["Small", "Medium"] #from PHB, Tweak?

table_names = [
"[customer_account]",
"[transactions]",
"[characters]",
"[spells_known]",
"[spells]",
"[inventory]",
"[items]",
"[weapons]",
"[armor]",
"[parties]",
"[encounters]",
"[monsters]"]

# #drop ifs: DROP TABLE IF EXISTS table CASCADE;
for i in range(len(table_names)):
    f.write("DROP TABLE IF EXISTS "+table_names[i]+";\n")
f.write("\n")
# # create tables
# customer_account(id, username, name, email, password, payment_rate)
create_ca = ("""CREATE TABLE """+table_names[0]+
                """
                    (id            INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    username       VARCHAR(100) NOT NULL,
                    name           VARCHAR(45) NOT NULL,
                    email          VARCHAR(45) NOT NULL,
                    password       VARCHAR(45) NOT NULL,
                    payment_rate   VARCHAR(10) NOT NULL
                    );""")

# transactions(id, fk customer_id, amount, date, status)
create_trans = ("CREATE TABLE "+table_names[1]+
                """
                    (id              INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    customer_id      INT NOT NULL,
                    amount           VARCHAR(45) NOT NULL,
                    _date            DATE NOT NULL,
                    status           VARCHAR(45) NOT NULL,
                    CONSTRAINT       customer_id_fk
                        FOREIGN KEY(customer_id)
                        REFERENCES customer_account(id)
                        ON DELETE CASCADE
                    );""")
# characters(id, name, fk party_id, fk customer_id, race, class, level, size)
create_chars = ("""CREATE TABLE """+table_names[2]+
                """
                    (id               INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name              VARCHAR(100) NOT NULL,
                    party_id          INT NOT NULL,
                    customer_id       INT NOT NULL,
                    race              VARCHAR(45) NOT NULL,
                    _class             VARCHAR(45) NOT NULL,
                    _level             VARCHAR(45) NOT NULL,
                    _size              VARCHAR(45) NOT NULL,
                    CONSTRAINT        party_id_fk
                        FOREIGN KEY(party_id)
                        REFERENCES parties(id)
                        ON DELETE CASCADE,
                    CONSTRAINT        customer_id_fk
                        FOREIGN KEY(customer_id)
                        REFERENCES customer_account(id)
                        ON DELETE CASCADE
                    );""")
# spells_known(fk character_id, fk spell_id)
#Tweak: character spells do not yet match level
create_spells_known = ("""CREATE TABLE """+table_names[3]+ #currently allow multiple of same spell to each character, need to make composite key
                """
                    (character_id     INT NOT NULL,
                    spell_id          INT NOT NULL,
                    CONSTRAINT        character_id_fk
                        FOREIGN KEY(character_id)
                        REFERENCES characters(id)
                        ON DELETE CASCADE,
                    CONSTRAINT       spell_id_fk
                        FOREIGN KEY(spell_id)
                        REFERENCES spells(id)
                        ON DELETE CASCADE,
                    PRIMARY KEY (character_id, spell_id)
                    );""")
# spells(id, name, spell level, description)
create_spells = ("""CREATE TABLE """+table_names[4]+
                """
                    (id                  INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name                 VARCHAR(100) UNIQUE NOT NULL,
                    spell_level          INT NOT NULL,
                    description          VARCHAR(300) NOT NULL
                    );""")
# inventory(fk character_id, fk item_id, quantity)
# #Tweak: need composite key
create_inv = ("""CREATE TABLE """+table_names[5]+
                """
                    (character_id      INT NOT NULL,
                    item_id           INT NOT NULL,
                    quantity          INT NOT NULL,
                    CHECK(quantity > 0),
                    CONSTRAINT        character_id_fk
                        FOREIGN KEY(character_id)
                        REFERENCES characters(id)
                        ON DELETE CASCADE,
                    CONSTRAINT        item_id_fk
                        FOREIGN KEY(item_id)
                        REFERENCES items(id)
                        ON DELETE CASCADE,
                    PRIMARY KEY (character_id,item_id)
                    );""")
# items(id, name, description), all info comes from excel files
create_items = ("""CREATE TABLE """+table_names[6]+
                """
                    (id                 INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name                VARCHAR(100) UNIQUE NOT NULL,
                    description         VARCHAR(300) NOT NULL
                    );""")
# # weapons(Id, name, description, properties, damage die, damage type)
create_weaps = ("""CREATE TABLE """+table_names[7]+
                """
                    (id                 INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name                VARCHAR(100) NOT NULL,
                    description         VARCHAR(300) NOT NULL,
                    properties          VARCHAR(100) NOT NULL,
                    damage_die          VARCHAR(10) NOT NULL,
                    damage_type         VARCHAR(45) NOT NULL
                    );""")
# # armor(Id, name, description, type, bonus, resistance)
create_arm = ("""CREATE TABLE """+table_names[8]+
                """
                    (id                INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name               VARCHAR(100) NOT NULL,
                    description        VARCHAR(300) NOT NULL,
                    _type               VARCHAR(45) NOT NULL,
                    bonus              VARCHAR(45),
                    resistance         VARCHAR(45)
                    );""")
# parties(id, name)
create_parties = ("""CREATE TABLE """+table_names[9]+
                """
                    (id                INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name               VARCHAR(100) NOT NULL
                    );""")
# encounters(fk party_id, fk monster_id, monster_deaths)
# Tweak: Needs composite key
create_enc = ("""CREATE TABLE """+table_names[10]+
                """
                    (party_id          INT NOT NULL,
                    monster_id         INT NOT NULL,
                    monster_deaths     INT NOT NULL,
                    CHECK(monster_deaths > 0),
                    CONSTRAINT        party_id_fk
                        FOREIGN KEY(party_id)
                        REFERENCES parties(id)
                        ON DELETE CASCADE,
                    CONSTRAINT        monster_id_fk
                        FOREIGN KEY(monster_id)
                        REFERENCES monsters(id)
                        ON DELETE CASCADE
                    );""")
# monsters(id, name, hit_points, exp_points)
create_monsters = ("""CREATE TABLE """+table_names[11]+
                """
                    (id                INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    name               VARCHAR(100) NOT NULL,
                    hit_points         INT,
                    exp_points         INT,
                    CHECK(hit_points > 0),
                    CHECK(exp_points > 0)
                    );""")

table_arr = [create_ca, create_trans, create_chars,
             create_spells_known, create_spells,
             create_inv, create_items, create_weaps,
             create_arm, create_parties, create_enc,
             create_monsters]

for i in range(len(table_arr)):
    f.write(table_arr[i]+"\r\n")
###end table creation
f.write("BEGIN;\r\n")


# # customer_account(id, username, name, email, password, payment_rate)
##generate username, name, email, passwords, payment rates from data files
##username generation: firstname[0]+lastname
num_customers = 2000 #must be > 1000, see instructions
for i in range(num_customers):
    p1 = misc_excel['A'+str(i+2)].value #for passwords
    p2 = misc_excel['B'+str(i+2)].value
    f.write("INSERT INTO customer_account (username, name, email, password, payment_rate) VALUES('"+
        #str(i)+", '"+ #id
        names_email_excel['A'+str(i+2)].value[0]+names_email_excel['B'+str(i+2)].value+"', '"+#username
        names_email_excel['A'+str(i+2)].value+" "+names_email_excel['B'+str(i+2)].value+"', '"+#name
        names_email_excel['C'+str(i+2)].value+"', '"+#email
        p1[:2]+p2[2:]+p1[2:]+p2[:2]+"', "+#password
        str(payment_rates[randint(0,3)])+#payment rate
        ");\r\n")

# # Transactions(id, customer_Id, amount, date, status)
num_trans = 600 #arbitrary
for i in range(num_trans):
    #Binary Decision Maker 3000:
    if bool(random.getrandbits(1)):
        status = "Pending"
    else:
        status = "Completed"

    f.write("INSERT INTO transactions (customer_id, amount, _date, status) VALUES("+#str(i)+", "+ #id
    str(randint(0,1999))+", '"+# customer_Id
    str(payment_rates[randint(1,3)]*randint(1,5))+"' , '"+# amount
    #^Tweak: very random, not related to individual subscriptions and ignores first pay_rate option
    misc_excel['E'+str(i+2)].value+"', '"+# date
    status +# status #Tweak: A trigger may change this behavior
    "');\r\n")
# # Characters(id, name, party_Id, customer_Id, race, class, level, size)
num_chars = 1500 #arbitrary

max_party_size = 7
party_size = 0
party_id = 0

for i in range(num_chars):
    B = names_email_excel['B'+str(i+2)].value
    C = misc_excel['C'+str(i+2)].value
    race = races[randint(0,len(races)-1)]
    if(race == "Halfling"):
        size = sizes[0]
    else:
        size = sizes[1]

    if(party_size == max_party_size):
        party_id+=1
        party_size=0
    party_size += 1

    f.write("INSERT INTO characters (name, party_Id, customer_Id, race, _class, _level, _size) VALUES('"+#str(i)+", '"+# id
    C+" of "+ B + "', "+# name
    str(party_id) + ", " +# party_Id
    str(randint(0, num_customers-1)) + ", '" +# customer_Id
    race + "', '" +# race
    classes[randint(0, len(classes)-1)] + "', " +# class
    str(randint(1,30)) + ", '" +# level #Tweak: Change to designed level cap.
    size +# size
    "');\r\n")



# # Spells_Known(character_id, spell_Id)
##Tweak?: allows spells to non-spell casters
num_spells_known = 400
num_spells = spells_excel.max_row-1 ##Tweak: If issues emerge: set to highest row number in excel -1
for i in range(num_spells_known):
    f.write("INSERT INTO spells_known VALUES("+
    str(randint(0, num_chars-1))+", "+ # character_id
    str(randint(0, num_spells-1))+# spell_Id #Tweak allow multiple of same spell
    ");\r\n")

# # Spells(Id, name, spell level, description)
for i in range(num_spells):
    f.write("INSERT INTO spells (name, spell_level, description) VALUES('"+#str(i)+", '"+# id
    spells_excel['A'+str(i+2)].value +"', "+ #name
    str(spells_excel['B'+str(i+2)].value) +", '"+ #spell level
    spells_excel['C'+str(i+2)].value + #description
    "');\r\n")

# # Inventory(character_id, item_Id, quantity)
num_inv = 300
num_items = items_excel.max_row-1

for i in range(num_inv):
    f.write("INSERT INTO inventory VALUES(" +
    str(randint(0, num_chars-1))+ "," +# character_id
    str(randint(0, num_items-1))+ "," +# item_Id
    str(randint(1, 25))+# quantity #Tweak?: arbitrary, not smart
    ");\r\n")

# # Parties(id, name)
num_parties = (party_id+1)
for i in range(num_parties):
    f.write("INSERT INTO parties (name) VALUES('"+#str(i)+", '"+# id
    misc_excel['A'+str(i+500)].value+" "+classes[randint(0,len(classes)-1)]+"s of "+misc_excel['B'+str(i+300)].value + # name #Tweak: Uniqueness not enforced
    "');\r\n")

# # Encounters(party_id, monster_Id, monster_deaths)
num_enc = randint(300, 500)
num_mons = monsters_excel.max_row-1
monster_deaths = 10 #Tweak: arbitrary
for i in range(num_enc):
    f.write("INSERT INTO encounters VALUES("+
    str(randint(0, num_parties-1))+", "+# party_id
    str(randint(0, num_mons - 1))+", "+# monster_Id
    str(randint(1, monster_deaths)) +# monster_deaths
    ");\r\n")

# # Monsters(Id, name, hit_points, exp_points)
for i in range(num_mons):
    f.write("INSERT INTO monsters (name, hit_points, exp_points) VALUES('"+#+#str(i)+", '"+# id
    monsters_excel['A'+str(i+2)].value +"', "+# name
    str(monsters_excel['B'+str(i+2)].value) +", "+# hit_points
    str(monsters_excel['C'+str(i+2)].value) +# exp_points
    ");\r\n")

# # Items(Id, name, description)
##item data stored in excel file
for i in range(num_items):
    f.write("INSERT INTO items (name, description) VALUES('"+#str(i)+", '"+# id
    items_excel['A'+str(i+2)].value+"', '" +# name
    items_excel['B'+str(i+2)].value +# description
    "');\r\n")

# # Weapons(Id, name, description, properties, damage die, damage type)
num_weapons = weapons_excel.max_row-1
for i in range(num_weapons):
    f.write("INSERT INTO weapons (name, description, properties, damage_die, damage_type) VALUES('"+#str(i)+", '"+# id
    weapons_excel['A'+str(i+2)].value+"', '" +# name
    weapons_excel['B'+str(i+2)].value+"', '" +# description
    weapons_excel['C'+str(i+2)].value+"', '" +# properties
    weapons_excel['D'+str(i+2)].value+"', '" +# damage die
    weapons_excel['E'+str(i+2)].value+ # damage type
    "');\r\n")

# # Armor(Id, name, description, type, bonus, resistance)
num_armor = armor_excel.max_row-1
for i in range(num_armor):
    resistance = str(armor_excel['E' + str(i+2)].value)
    if(resistance =='None'):
        resistance = 'Null'
    else:
        resistance = "'"+resistance+"'"

    f.write("INSERT INTO armor (name, description, _type, bonus, resistance) VALUES('"+#str(i)+", '" +# id
    armor_excel['A' + str(i+2)].value+"', '" +# name
    armor_excel['B' + str(i+2)].value+"', '" +# description
    armor_excel['C' + str(i+2)].value+"', '" +# type
    str(armor_excel['D' + str(i + 2)].value)+"', " +# bonus
    resistance+ # resistance
    ");\r\n")
f.write("COMMIT;")
f.close()