#Creates a basic 'table.sql' and 'inserts.sql' files
#  IMPORTANT: This will overwrite any file named 'table.sql' and 'inserts.sql'!

import openpyxl #!!required: pip install openyxl
from random import randint
import random

#f = open("table.sql","w")
g = open("inserts.sql","w")

#Excel I/O
wb_armor = openpyxl.load_workbook('armor.xlsx')
armor_excel = wb_armor.active
wb_items = openpyxl.load_workbook('items.xlsx')
items_excel = wb_items.active
wb_weapons = openpyxl.load_workbook('weapons.xlsx')
weapons_excel = wb_weapons.active
wb_names_email = openpyxl.load_workbook('first_last_email.xlsx')
names_email_excel = wb_names_email.active
wb_misc = openpyxl.load_workbook('misc.xlsx')
misc_excel = wb_misc.active
wb_spells = openpyxl.load_workbook('spells.xlsx')
spells_excel = wb_spells.active
wb_monsters = openpyxl.load_workbook('monsters.xlsx')
monsters_excel = wb_monsters.active

payment_rates = [0, 15] #Tweak: Change to design.
hit_die = [4,6,8,10,12,20] #Tweak: possibly unused
races = ["Elf","Dwarf","Human","Halfling","Orc", "Dragonborn", "Tiefling", "Half-Elf", "Half-Orc"] # from PHB --Tweak
classes = ["Paladin","Cleric","Wizard", "Rogue","Ranger", "Monk", "Sorcerer", "Druid", "Barbarian", "Bard"] # from PHB --Tweak
sizes = ["Small", "Medium"] #from PHB, Tweak?

# # customer_account(id, username, name, email, password, payment_rate)
##generate username, name, email, passwords, payment rates from data files
##username generation: firstname[0]+lastname
num_customers = 2000 #must be > 1000, see instructions
for i in range(num_customers):
    fn = names_email_excel['A'+str(i+2)].value#first name
    ln = names_email_excel['B'+str(i+2)].value#last name
    p1 = misc_excel['A'+str(i+2)].value #for passwords
    p2 = misc_excel['B'+str(i+2)].value
    g.write("INSERT INTO customer_account (username, name, email, password, payment_rate) VALUES('"+
        fn[0]+ln+"', '"+#username
        fn+" "+ln+"', '"+#name
        names_email_excel['C'+str(i+2)].value+"', '"+#email
        p1[:2]+p2[2:]+p1[2:]+p2[:2]+"', "+#password
        str(payment_rates[randint(0,1)])+#payment rate
        ");\r\n")

# # Transactions(id, customer_Id, amount, date, status)
num_trans = 1200 #arbitrary
for i in range(num_trans):
    #Binary Decision Maker 3000:
    if bool(random.getrandbits(1)):
        status = "Pending"
    else:
        status = "Completed"

    g.write("INSERT INTO transactions (customer_id, amount, _date, status) VALUES("+#id
    str(randint(1,num_customers))+", '"+# customer_Id
    str(payment_rates[1]*randint(1,5))+"' , '"+# amount
    #^Tweak: very random, not related to individual subscriptions and ignores first pay_rate option
    misc_excel['E'+str(i+2)].value+"', '"+# date
    status +# status #Tweak: A trigger may change this behavior
    "');\r\n")

# # Parties(id, name)
num_parties = (int(4000/7)+1)
for i in range(num_parties):
    g.write("INSERT INTO parties (name) VALUES('"+
    misc_excel['A'+str(i+500)].value+" "+classes[randint(0,len(classes)-1)]+"s of "+misc_excel['F'+str(i+300)].value + # name #Tweak: Uniqueness not enforced
    "');\r\n")

# # Characters(id, name, party_Id, customer_Id, race, class, level, size)
num_chars = 2000 #arbitrary
max_party_size = 7
party_size = 0
party_id = 1

for i in range(num_chars*2):
    B = names_email_excel['B'+str(randint(2,num_chars-2))].value
    C = misc_excel['C'+str(randint(2,num_chars-2))].value
    race = races[randint(0,len(races)-1)]
    if(race == "Halfling"):
        size = sizes[0]
    else:
        size = sizes[1]

    if(party_size == max_party_size):
        party_id+=1
        party_size=0
    party_size += 1

    g.write("INSERT INTO characters (name, party_Id, customer_Id, race, _class, _level, _size) VALUES('"+#str(i)+", '"+# id
    C+" of "+ B + "', "+# name
    str(party_id) + ", " +# party_Id
    str(randint(1, num_customers-1)) + ", '" +# customer_Id
    race + "', '" +# race
    classes[randint(0, len(classes)-1)] + "', " +# class
    str(randint(1,30)) + ", '" +# level #Tweak: Change to designed level cap.
    size +# size
    "');\r\n")

# # Spells(Id, name, spell level, description)
num_spells = spells_excel.max_row-1 ##Tweak: If issues emerge: set to highest row number in excel -1
for i in range(num_spells):
    g.write("INSERT INTO spells (name, spell_level, description) VALUES('"+
    spells_excel['A'+str(i+2)].value +"', "+ #name
    str(spells_excel['B'+str(i+2)].value) +", '"+ #spell level
    spells_excel['C'+str(i+2)].value + #description
    "');\r\n")

# # Spells_Known(character_id, spell_Id)
##Tweak?: allows spells to non-spell casters
num_spells_known = 3200
for i in range(1,num_spells_known+1):
    g.write("INSERT INTO spells_known VALUES("+
    str(i)+", "+ # character_id
    str(randint(1, 10))+# spell_Id 
    ");\r\n")
    g.write("INSERT INTO spells_known VALUES("+
    str(i)+", "+ # character_id
    str(randint(11, 21))+# spell_Id
    ");\r\n")
    g.write("INSERT INTO spells_known VALUES("+
    str(i)+", "+ # character_id
    str(randint(22, 31))+# spell_Id
    ");\r\n")
    g.write("INSERT INTO spells_known VALUES("+
    str(i)+", "+ # character_id
    str(randint(32, num_spells-1))+# spell_Id 
    ");\r\n")

# # Monsters(Id, name, hit_points, exp_points)
num_mons = monsters_excel.max_row-1
for i in range(num_mons):
    g.write("INSERT INTO monsters (name, hit_points, exp_points) VALUES('"+# id
    monsters_excel['A'+str(i+2)].value +"', "+# name
    str(monsters_excel['B'+str(i+2)].value) +", "+# hit_points
    str(monsters_excel['C'+str(i+2)].value) +# exp_points
    ");\r\n")

# # Encounters(party_id, monster_Id, monster_deaths)
num_enc = 800 #arbitrary
monster_deaths = 10 #Tweak: arbitrary
for i in range(num_enc):
    g.write("INSERT INTO encounters VALUES("+
    str(randint(1, num_parties-1))+", "+# party_id
    str(randint(1, num_mons - 1))+", "+# monster_Id
    str(randint(1, monster_deaths)) +# monster_deaths
    ");\r\n")

# # Items(Id, name, description)
##item data stored in excel file
num_items = items_excel.max_row-1
num_weapons = weapons_excel.max_row-1
num_armor = armor_excel.max_row-1

for i in range(1, num_items+1):
    g.write("INSERT INTO items (name, description) VALUES('"+
    items_excel['A'+str(i+1)].value+"', '" +# name
    items_excel['B'+str(i+1)].value +# description
    "');\r\n")
for i in range(1, num_weapons+1):
    g.write("INSERT INTO items (name, description) VALUES('"+
    weapons_excel['A'+str(i+1)].value+"', '" +# name
    weapons_excel['B'+str(i+1)].value +# description
    "');\r\n")
for i in range(1, num_armor+1):
    g.write("INSERT INTO items (name, description) VALUES('"+
    armor_excel['A'+str(i+1)].value+"', '" +# name
    armor_excel['B'+str(i+1)].value +# description
    "');\r\n")

# # Weapons(Id, name, description, properties, damage die, damage type)
max_weap_id = num_weapons+num_items+1
j = 0
for i in range(num_items+1,max_weap_id):
    g.write("INSERT INTO weapons (id, name, description, properties, damage_die, damage_type) VALUES("+
    str(i)+", '"+# id
    weapons_excel['A'+str(j+2)].value+"', '" +# name
    weapons_excel['B'+str(j+2)].value+"', '" +# description
    weapons_excel['C'+str(j+2)].value+"', " +# properties
    str(hit_die[randint(0,len(hit_die)-1)])+", '" +#weapons_excel['D'+str(j+2)].value+"', '" +# damage die
    weapons_excel['E'+str(j+2)].value+ # damage type
    "');\r\n")
    j+=1
	
# # Armor(Id, name, description, type, bonus, resistance)
max_armor_id = num_armor+num_items+num_weapons
j = 0
for i in range(max_weap_id,max_armor_id):
    resistance = armor_excel['E' + str(j+2)].value
    # if(resistance =='Normal'):
    #     resistance = 'Null'
    # else:
    #     resistance = "'"+resistance+"'"

    g.write("INSERT INTO armor (id, name, description, _type, bonus, resistance) VALUES("+
    str(i) + ", '" +  # id
    armor_excel['A' + str(j+2)].value+"', '" +# name
    armor_excel['B' + str(j+2)].value+"', '" +# description
    armor_excel['C' + str(j+2)].value+"', '" +# type
    armor_excel['D' + str(j+2)].value+"', '" +# bonus
    resistance+ # resistance
    "');\r\n")
    j+=1

# # Inventory(character_id, item_Id, quantity)
num_inv = 300
for i in range(1, (num_chars*2)+1):
    g.write("INSERT INTO inventory VALUES(" +
    str(i)+ "," +# character_id
    str(randint(1, 19))+ "," +# item_Id
    str(randint(1, 10))+# quantity #Tweak?: arbitrary, not smart
    ");\r\n")
    g.write("INSERT INTO inventory VALUES(" +
    str(i)+ "," +# character_id
    str(randint(20, 38))+ "," +# item_Id
    str(randint(1, 10))+# quantity #Tweak?: arbitrary, not smart
    ");\r\n")
    g.write("INSERT INTO inventory VALUES(" +
    str(i)+ "," +# character_id
    str(randint(39, 57))+ "," +# item_Id
    str(randint(1, 10))+# quantity #Tweak?: arbitrary, not smart
    ");\r\n")
    g.write("INSERT INTO inventory VALUES(" +
    str(i)+ "," +# character_id
    str(randint(58, 77))+ "," +# item_Id
    str(randint(1, 10))+# quantity #Tweak?: arbitrary, not smart
    ");\r\n")
	
g.close()
